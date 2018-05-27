from flask import Flask, jsonify,request
import googlemaps
from datetime import datetime
import math
import os
import numpy as np
import json
import requests
import operator
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.interpolate import spline

from selenium import webdriver
import urllib

# def get_js():
#     # f = open("D:/WorkSpace/MyWorkSpace/jsdemo/js/des_rsa.js",'r',encoding='UTF-8')
#     f = open("/Users/wyj/Desktop/comp9321/untitled2/a.js", 'r', encoding='UTF-8')
#     line = f.readline()
#     htmlstr = ''
#     while line:
#         htmlstr = htmlstr + line
#         line = f.readline()
#     return htmlstr
# print(os.getcwd())

# jsstr = get_js()
# ctx = execjs.compile(jsstr)
# print(ctx.call('enString', '123456'))

# x1,y1 = -33.5297025187805,148.7673155965424
#
# x2, y2 = -33.529546006456854,148.76623734851842
#
# x0,y0 = -33.529089883212144,148.76858696361546
# d = abs((y2 - y1) * x0 +(x1 - x2) * y0 + ((x2 * y1) -(x1 * y2)))/(math.sqrt(pow(y2 - y1, 2) + pow(x1 - x2, 2)))
# print(d)

#
# url1 = "https://docs.google.com/spreadsheets/d/1tHCxouhyM4edDvF60VG7nzs5QxID3ADwr3DGJh71qFg/export?format=xlsx&id=1tHCxouhyM4edDvF60VG7nzs5QxID3ADwr3DGJh71qFg"
# r = requests.get(url1)
# lga_set = set()
# # lga_mapping = {}
# with open("temp_postcode.xlsx", "wb") as code:
#     code.write(r.content)
# wb = load_workbook(filename=r'temp_postcode.xlsx')
# sheet = wb.get_sheet_by_name('lga_postcode_mappings')
# for i in range(2,1783):
#     lga_set.add(sheet.cell(row=i, column=2).value)
    # lga_mapping[int(sheet.cell(row=i, column=3).value)] = sheet.cell(row=i, column=2).value


# lga_list = {}
# x = open('/Users/wyj/Desktop/geoserver-GetFeature.json','r')
# load_dict = json.load(x)
# #
#
# for i in range(len(load_dict['features'])):
#     if load_dict['features'][i]['properties']['nsw_lga__3'] != 'UNINCORPORATED':
#         t = ''.join([x for x in load_dict['features'][i]['properties']['nsw_lga__3'].lower() if x.isalpha()])
#
#         temp_list = load_dict['features'][i]['geometry']['coordinates'][0][0]
#
#         j =0
#         while j < len(temp_list)-2:
#             x1,y1 = temp_list[j]
#             x2, y2 = temp_list[j+1]
#             x0,y0 = temp_list[j+2]
#             if (abs((y2 - y1) * x0 +(x1 - x2) * y0 + ((x2 * y1) -(x1 * y2)))/(math.sqrt(pow(y2 - y1, 2) + pow(x1 - x2, 2)))) <0.00001:
#
#                 temp_list.pop(j+1)
#             else:
#
#
#                 j += 1
#         # j = 0
#         # while j < len(temp_list) - 2:
#         #     x1, y1 = temp_list[j]
#         #     x2, y2 = temp_list[j + 1]
#         #     x0, y0 = temp_list[j + 2]
#         #     if (abs((y2 - y1) * x0 + (x1 - x2) * y0 + ((x2 * y1) - (x1 * y2))) / (
#         #     math.sqrt(pow(y2 - y1, 2) + pow(x1 - x2, 2)))) < 0.0001:
#         #         temp_list.pop(j + 1)
#         #
#         #     j += 1
#
#
#         lga_list[t] = temp_list


#
#

# with open("/Users/wyj/Desktop/lgalist5.json","w") as f:
#
#     json.dump(lga_list, f)
x = open('/Users/wyj/Desktop/final_lga.json','r')
lga_dic = json.load(x)
lga_set = set()
for key in lga_dic:
    lga_set.add(key)
print(lga_set)
app = Flask(__name__)


@app.route('/get_all_crimedata', methods=['GET'])
def get_all_crimedata():
    response = requests.get("http://localhost:5002/nsw_crime_data", params=None)
    # print("statistics:", response.json())
    load_dict = response.json()
    json = {}
    sorted_x = sorted(load_dict['entry'], key=lambda k: k['average'])
    # print(sorted_x)
    for i in range(len(sorted_x)):
        if sorted_x[i]['lga_name'] in lga_dic:

            # coordinate = lga_dic[sorted_x[i]['lga_name']]
            data = {
                    'average':  sorted_x[i]['average'],
                    'rank': i}
            json[sorted_x[i]['lga_name']] = data
        else:
            print(sorted_x[i]['lga_name'])

    return jsonify(json)

@app.route('/get_one_crimedata/<string:lga_id>', methods=['GET'])
def get_one_crimedata(lga_id):
    lga_name = lga_id
    lga_name = ''.join([x for x in lga_name.lower() if x.isalpha()])
    print(lga_name)
    response = requests.get("http://localhost:5002/nsw_crime_data/"+ lga_name, params=None)
    print("statistics:", response.json())
    load_dict = response.json()
    json = {}

    # coordinate = lga_dic[lga_name]
    temp_year_data = load_dict['year_data']
    X_ = []
    Y_ = []
    for key_  in temp_year_data:
        X_.append(int(key_))
        Y_.append(float(temp_year_data[key_]))
    X_ = np.array(X_)
    Y_ = np.array(Y_)
    xnew = np.linspace(min(X_), max(X_), 300)
    t = spline(X_, Y_, xnew)
    plt.plot(xnew, t)
    plt.xlabel("Year")
    plt.ylabel("Crimes")
    plt.title(lga_name+" Recorded Crime Statistics")
    plt.savefig("crimes.png")
    path  = os.getcwd()+"/crimes.png"
    print(path)
    data = {'year_data': load_dict['year_data'],
            'average': load_dict['average'],
            'path':path}
    json[load_dict['lga_name']] = data

    return jsonify(json)


@app.route('/get_all_rent', methods=['GET'])
def get_all_rent():
    response = requests.get("http://localhost:5001/nsw_rent_data", params=None)
    print("statistics:", response.json())
    load_dict = response.json()
    json = {}

    temp_dict = {}
    for x in range(len(load_dict['entry'])):
        temp_list = []
        if load_dict['entry'][x]['one_bed'] != 'null':
            temp_list.append(float(load_dict['entry'][x]['one_bed']))
        if load_dict['entry'][x]['one_bed'] != 'null':
            temp_list.append(float(load_dict['entry'][x]['two_bed'])/2)
        if load_dict['entry'][x]['one_bed'] != 'null':
            temp_list.append(float(load_dict['entry'][x]['three_bed'])/3)
        if load_dict['entry'][x]['one_bed'] != 'null':
            temp_list.append(float(load_dict['entry'][x]['four_bed'])/4)
        if len(temp_list)!= 0:
            temp_dict[load_dict['entry'][x]['lga_name']] = temp_list
    # print(temp_dict.items())
    sort_x = sorted(temp_dict.items(),key=lambda a: float(sum(a[1]))/len(a[1]))
    print(sort_x)
    rank_list = {}
    for ii in range(len(sort_x)):

        rank_list[sort_x[ii][0]] = ii
    print(rank_list)
    # sorted_x = sorted(load_dict['entry'], key=lambda k: (float(k['one_bed'])+float(k['two_bed'])/2+float(k['three_bed'])/3+float(k['four_bed'])/4))
    sorted_x = load_dict['entry']

    for i in range(len(sorted_x)):
        if sorted_x[i]['lga_name'] in lga_dic:
            if sorted_x[i]['lga_name'] in rank_list:
                rank = rank_list[sorted_x[i]['lga_name']]
            else:
                rank = len(sorted_x)
            # coordinate = lga_dic[sorted_x[i]['lga_name']]
            data = {
                    'rank': rank }
            json[sorted_x[i]['lga_name']] = data
        else:
            print(sorted_x[i]['lga_name'])


    return jsonify(json)
@app.route('/get_one_rent/<string:lga_id>', methods=['GET'])
def get_one_rent(lga_id):
    lga_name = lga_id

    lga_name = ''.join([x for x in lga_name.lower() if x.isalpha()])
    response = requests.get("http://localhost:5001/nsw_rent_data/" + lga_name, params=None)
    print("statistics:", response.json())
    load_dict = response.json()
    json = {}

    # coordinate = lga_dic[lga_name]
    name_list = ['One Bed', 'Two Bed', 'Three Bed', 'Four Bed']
    num_list = [load_dict['one_bed'], load_dict['two_bed'], load_dict['three_bed'], load_dict['four_bed']]
    num_list1 = [load_dict['annual_rate_one_bed'], load_dict['annual_rate_two_bed'], load_dict['annual_rate_three_bed'], load_dict['annual_rate_four_bed']]
    x = list(range(len(num_list)))
    total_width, n = 0.8, 2
    width = total_width / n

    plt.bar(x, num_list, width=width, label='Price', fc='y')
    for i in range(len(x)):
        x[i] = x[i] + width
    plt.bar(x, num_list1, width=width, label='annual rate', tick_label=name_list, fc='r')

    plt.ylabel("Price")
    plt.title(lga_name+" Recorded RENT Statistics")
    plt.savefig("rents.png")
    path = os.getcwd()+"/rents.png"
    print(path)
    data = {
            'path': path}
    json[load_dict['lga_name']] =data

    return jsonify(json)

@app.route('/get_all_sales', methods=['GET'])
def get_all_sales():
    response = requests.get("http://localhost:5001/nsw_sales_data", params=None)
    # print("statistics:", response.json())
    load_dict = response.json()
    print(type(response))
    json = {}
    sorted_x = sorted(load_dict['entry'],
                      key=lambda k: k['median'])
    for i in range(len(sorted_x)):
        if sorted_x[i]['lga_name'] in lga_dic:

            # coordinate = lga_dic[sorted_x[i]['lga_name']]
            data = {
                'rank': i}
            json[sorted_x[i]['lga_name']] = data
        else:
            print(sorted_x[i]['lga_name'])

    return jsonify(json)
@app.route('/get_one_sale/<string:lga_id>', methods=['GET'])
def get_one_sale(lga_id):
    lga_name = lga_id
    lga_name = ''.join([x for x in lga_name.lower() if x.isalpha()])
    print(lga_name)
    response = requests.get("http://localhost:5001/nsw_sales_data/" + lga_name, params=None)
    print("statistics:", response.json())
    load_dict = response.json()
    json = {}

    # coordinate = lga_dic[lga_name]

    data = {'median': load_dict['median'],
            'annual_rate_median': load_dict['annual_rate_median'],
            }
    json[load_dict['lga_name']] = data

    return jsonify(json)

@app.route('/get_all_coordinates', methods=['GET'])
def get_all_coordinates():
    coordinate = lga_dic
    return jsonify(coordinate)

# plt.legend()
# plt.show()

@app.route('/get_all_rank', methods=['GET'])
def get_all_set():
    total_rank = {}
    # sale_rank = get_all_sales()
    response = requests.get("http://localhost:5000/get_all_sales", params=None)
    # print("type",type(sale_rank))
    sale_dic = response.json()
    # print('kkkkkk', sale_dic)
    response = requests.get("http://localhost:5000/get_all_rent", params=None)
    # rent_rank = get_all_rent()
    rent_dic = response.json()
    response = requests.get("http://localhost:5000/get_all_crimedata", params=None)
    # crime_rank = get_all_crimedata()
    crime_dic = response.json()
    print(sale_dic)
    for ele in lga_set:
        lga_name = ''.join([x for x in ele.lower() if x.isalpha()])

        all_rank = []
        if lga_name in sale_dic:
            all_rank.append(sale_dic[lga_name]['rank'])
        if lga_name in rent_dic:
            all_rank.append(rent_dic[lga_name]['rank'])
        if lga_name in crime_dic:
            all_rank.append(crime_dic[lga_name]['rank'])
        if len(all_rank) !=0:
            total_rank[lga_name] = sum(all_rank)/len(all_rank)
        else:
            total_rank[lga_name] = len(lga_set)
    rank_list = {}
    sort_x = sorted(total_rank.items(), key=lambda a: a[1])
    for ii in range(len(sort_x)):
        rank_list[sort_x[ii][0]] = ii
    return jsonify(rank_list)

if __name__ == '__main__':
    app.run()
